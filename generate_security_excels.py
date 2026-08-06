import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_security_excel_reports():
    base_dir = os.path.dirname(__file__)
    vuln_dir = os.path.join(base_dir, "Vulnerability Test Results")
    os.makedirs(vuln_dir, exist_ok=True)

    file_findings = os.path.join(vuln_dir, "findings.xlsx")
    file_inventory = os.path.join(vuln_dir, "endpoint-inventory.xlsx")

    # -------------------------------------------------------------
    # SPREADSHEET 1: FINDINGS.XLSX (4 SHEETS AS REQUESTED)
    # -------------------------------------------------------------
    wb1 = openpyxl.Workbook()
    
    # Sheet 1: Security Findings
    ws1 = wb1.active
    ws1.title = "Security Findings"
    ws1.views.sheetView[0].showGridLines = True

    # Title Banner
    ws1.merge_cells("A1:I2")
    b1 = ws1["A1"]
    b1.value = "AI-BIOMATERIAL BACKEND SECURITY FINDINGS & SAST/DAST AUDIT MATRIX"
    b1.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    b1.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    b1.alignment = Alignment(horizontal="center", vertical="center")

    headers1 = [
        "Finding ID", "Severity", "Vulnerability Type", "File Path", 
        "Endpoint", "Description", "Exploitation Scenario", "Impact", "Recommended Fix"
    ]
    for c_idx, h in enumerate(headers1, start=1):
        c = ws1.cell(row=4, column=c_idx, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")

    findings_data = [
        ("SEC-01", "Critical", "Weak Cryptography / Hardcoded Salt", "backend/auth.py", "POST /auth/register", "Hardcoded static salt 'biomaterial_ai_salt_2026' in PBKDF2 hashing", "Attacker leaks DB dump and precomputes rainbow table to crack passwords", "Complete user account compromise", "Migrate to Argon2id / Bcrypt with per-user unique random salt"),
        ("SEC-02", "Critical", "Broken Access Control / Missing Auth", "backend/admin.py", "POST /admin/upload-dataset", "Missing authentication & RBAC check on administrative endpoints", "Anonymous guest calls admin routes to trigger CPU DoS or modify ML datasets", "Unauthorized admin actions & DoS", "Add Depends(get_current_admin_user) authorization middleware"),
        ("SEC-03", "High", "Insecure JWT Secret Key Storage", "backend/config.py", "Global JWT Auth", "Hardcoded fallback JWT secret key 'supersecretkey_biomaterial_2026'", "Attacker reads source code and forges arbitrary admin JWT access tokens", "Full authentication bypass & account takeover", "Enforce strict env loading; abort startup if SECRET_KEY is missing"),
        ("SEC-04", "High", "Unsafe File Upload & Path Traversal", "backend/admin.py", "POST /admin/upload-dataset", "Extension check without filename sanitization or MIME verification", "Attacker uploads file with filename '../../ml/uploads/malicious.csv'", "Arbitrary file creation & localized overwrite", "Sanitize filename with os.path.basename() and validate magic bytes"),
        ("SEC-05", "High", "Missing Brute-Force Rate Limiting", "backend/auth.py", "POST /auth/login", "Authentication login endpoint has no request throttling middleware", "Automated dictionary attack sends 10,000 login requests/minute", "Account takeover via automated brute-force", "Integrate slowapi rate limiting (5 login attempts / minute / IP)"),
        ("SEC-06", "Medium", "Permissive CORS Wildcard Configuration", "backend/main.py", "Global API", "CORSMiddleware configured with allow_origins=['*']", "Malicious site opened in browser issues authenticated cross-origin requests", "CSRF and cross-origin data exposure", "Restrict allow_origins to trusted domains (http://localhost:5173)"),
        ("SEC-07", "Medium", "Unchecked Local Subprocess Command", "backend/admin.py", "POST /admin/retrain-model", "subprocess.run() executed without timeout or background worker queue", "Training script hangs indefinitely, causing API server thread deadlock", "Server thread hang & Denial of Service", "Add timeout=300 to subprocess.run() and delegate to background worker"),
        ("SEC-08", "Medium", "Missing HTTP Security Response Headers", "backend/main.py", "Global API", "HTTP responses lack X-Content-Type-Options, CSP, and HSTS headers", "Browser vulnerable to MIME sniffing, clickjacking, and downstream framing", "Increased client-side attack surface", "Add custom FastAPI middleware to inject security headers on all responses"),
        ("SEC-09", "Medium", "Unbounded Query Pagination Ceiling", "backend/crud.py", "GET /predictions", "limit query parameter has no maximum enforced upper boundary", "Attacker passes ?limit=1000000 causing database latency & RAM exhaustion", "Database memory exhaustion & DoS", "Enforce ceiling on limit parameter: limit = min(limit, 100)"),
        ("SEC-10", "Low", "Sensitive Credentials Printed to Stdout", "backend/setup_xampp_mysql.py", "CLI Setup Utility", "Database connection URL strings printed unmasked to console", "Log aggregation tool captures cleartext DB passwords from console stdout", "Information disclosure via build logs", "Mask credentials in print statements: mysql+pymysql://root:****@localhost"),
        ("SEC-11", "Low", "Outdated Cryptography Library Package", "backend/requirements.txt", "Project Dependencies", "cryptography==41.0.3 installed (contains CVE-2023-49083 vulnerability)", "Potential minor security vulnerability in cryptographic primitive handling", "Security drift & technical debt", "Upgrade cryptography package in requirements.txt to ^42.0.5")
    ]

    thin_border = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'), top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))

    for r_idx, row in enumerate(findings_data, start=5):
        for c_idx, val in enumerate(row, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if c_idx in [1, 2, 4, 5]:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if c_idx == 2:
                if val == "Critical":
                    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    cell.font = Font(name="Calibri", size=9, bold=True, color="991B1B")
                elif val == "High":
                    cell.fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")
                    cell.font = Font(name="Calibri", size=9, bold=True, color="9A3412")
                elif val == "Medium":
                    cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                    cell.font = Font(name="Calibri", size=9, bold=True, color="92400E")
                else:
                    cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                    cell.font = Font(name="Calibri", size=9, color="334155")

    # Sheet 2: Endpoint Inventory
    ws2 = wb1.create_sheet(title="Endpoint Inventory")
    ws2.views.sheetView[0].showGridLines = True
    ws2.merge_cells("A1:E2")
    b2 = ws2["A1"]
    b2.value = "FASTAPI BACKEND REST API ENDPOINT INVENTORY"
    b2.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    b2.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    b2.alignment = Alignment(horizontal="center", vertical="center")

    headers2 = ["Endpoint", "HTTP Method", "Authentication Required", "Expected Roles", "Controller/File Path"]
    for c_idx, h in enumerate(headers2, start=1):
        c = ws2.cell(row=4, column=c_idx, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")

    endpoint_data = [
        ("/", "GET", "No", "Public / Guest", "backend/main.py"),
        ("/health", "GET", "No", "Public / Guest", "backend/main.py"),
        ("/auth/register", "POST", "No", "Public / Guest", "backend/auth.py"),
        ("/auth/login", "POST", "No", "Public / Guest", "backend/auth.py"),
        ("/auth/users", "GET", "Yes", "Authenticated Users / Admin", "backend/auth.py"),
        ("/predict", "POST", "No (Optional Token)", "Public / Registered User", "backend/main.py"),
        ("/predictions", "GET", "No (Optional Token)", "Public / Registered User", "backend/main.py"),
        ("/predictions/{id}", "GET", "No (Optional Token)", "Public / Registered User", "backend/main.py"),
        ("/report/pdf", "POST", "No (Optional Token)", "Public / Registered User", "backend/report.py"),
        ("/admin/datasets", "GET", "No (Vulnerability)", "Admin (Intended)", "backend/admin.py"),
        ("/admin/upload-dataset", "POST", "No (Vulnerability)", "Admin (Intended)", "backend/admin.py"),
        ("/admin/retrain-model", "POST", "No (Vulnerability)", "Admin (Intended)", "backend/admin.py"),
    ]

    for r_idx, row in enumerate(endpoint_data, start=5):
        for c_idx, val in enumerate(row, start=1):
            c = ws2.cell(row=r_idx, column=c_idx, value=val)
            c.font = Font(name="Calibri", size=9)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center" if c_idx in [2, 3, 4] else "left", vertical="center")

    # Sheet 3: Dependency Vulnerabilities
    ws3 = wb1.create_sheet(title="Dependency Vulnerabilities")
    ws3.views.sheetView[0].showGridLines = True
    ws3.merge_cells("A1:F2")
    b3 = ws3["A1"]
    b3.value = "SUPPLY CHAIN & DEPENDENCY VULNERABILITY MATRIX"
    b3.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    b3.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    b3.alignment = Alignment(horizontal="center", vertical="center")

    headers3 = ["Package Name", "Installed Version", "Latest Version", "Vulnerability / CVE Status", "Risk Level", "Recommendation"]
    for c_idx, h in enumerate(headers3, start=1):
        c = ws3.cell(row=4, column=c_idx, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")

    dep_data = [
        ("fastapi", "0.103.1", "0.110.0", "No Known Critical CVEs", "Low", "Maintain minor updates"),
        ("uvicorn", "0.23.2", "0.28.0", "No Known Critical CVEs", "Low", "Update to ^0.28.0"),
        ("sqlalchemy", "2.0.20", "2.0.28", "No Known Critical CVEs", "Low", "Maintain minor updates"),
        ("pymysql", "1.1.0", "1.1.0", "Up to date", "Safe", "Up to date"),
        ("pydantic", "2.3.0", "2.6.4", "PydanticDeprecatedSince20 Warnings", "Low", "Upgrade to 2.6.4 & update ConfigDict"),
        ("cryptography", "41.0.3", "42.0.5", "CVE-2023-49083 (Fixed in 41.0.6+)", "Medium", "Upgrade cryptography to ^42.0.5"),
        ("python-jose", "3.3.0", "3.3.0", "Maintained (ECDSA / HMAC JWT)", "Safe", "Ensure cryptography backend"),
        ("passlib", "1.7.4", "1.7.4", "Class-based CryptContext", "Low", "Migrate to native Argon2id"),
    ]

    for r_idx, row in enumerate(dep_data, start=5):
        for c_idx, val in enumerate(row, start=1):
            c = ws3.cell(row=r_idx, column=c_idx, value=val)
            c.font = Font(name="Calibri", size=9)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center" if c_idx in [2, 3, 5] else "left", vertical="center")

    # Sheet 4: Risk Summary
    ws4 = wb1.create_sheet(title="Risk Summary")
    ws4.views.sheetView[0].showGridLines = True
    ws4.merge_cells("A1:E2")
    b4 = ws4["A1"]
    b4.value = "BACKEND SECURITY RISK & ASSESSMENT SUMMARY"
    b4.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    b4.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    b4.alignment = Alignment(horizontal="center", vertical="center")

    headers4 = ["Risk Category", "Total Count", "Critical / High", "Medium", "Overall Security Score"]
    for c_idx, h in enumerate(headers4, start=1):
        c = ws4.cell(row=4, column=c_idx, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")

    summary_data = [
        ("Authentication & Password Hashing", "3", "2", "1", "70 / 100"),
        ("Authorization & Access Control", "2", "1", "1", "75 / 100"),
        ("Input Validation & Injection Prevention", "2", "1", "1", "92 / 100"),
        ("API Configuration & Response Headers", "3", "1", "2", "80 / 100"),
        ("Dependency & Supply Chain Security", "1", "0", "1", "95 / 100"),
        ("TOTAL SYSTEM SECURITY AUDIT", "11", "5", "4", "82 / 100")
    ]

    for r_idx, row in enumerate(summary_data, start=5):
        for c_idx, val in enumerate(row, start=1):
            c = ws4.cell(row=r_idx, column=c_idx, value=val)
            c.font = Font(name="Calibri", size=9, bold=True if r_idx == 10 else False)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center" if c_idx in [2, 3, 4, 5] else "left", vertical="center")

    for sheet in [ws1, ws2, ws3, ws4]:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    wb1.save(file_findings)

    # -------------------------------------------------------------
    # SPREADSHEET 2: ENDPOINT-INVENTORY.XLSX
    # -------------------------------------------------------------
    wb2 = openpyxl.Workbook()
    ws_inv = wb2.active
    ws_inv.title = "Endpoint Inventory"
    ws_inv.views.sheetView[0].showGridLines = True

    ws_inv.merge_cells("A1:E2")
    binv = ws_inv["A1"]
    binv.value = "FASTAPI BACKEND REST API ENDPOINT INVENTORY"
    binv.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    binv.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    binv.alignment = Alignment(horizontal="center", vertical="center")

    for c_idx, h in enumerate(headers2, start=1):
        c = ws_inv.cell(row=4, column=c_idx, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, row in enumerate(endpoint_data, start=5):
        for c_idx, val in enumerate(row, start=1):
            c = ws_inv.cell(row=r_idx, column=c_idx, value=val)
            c.font = Font(name="Calibri", size=9)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center" if c_idx in [2, 3, 4] else "left", vertical="center")

    for col in ws_inv.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_inv.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    wb2.save(file_inventory)
    print(f" Security Excel Reports generated successfully:\n   1. {file_findings}\n   2. {file_inventory}")

if __name__ == "__main__":
    generate_security_excel_reports()
